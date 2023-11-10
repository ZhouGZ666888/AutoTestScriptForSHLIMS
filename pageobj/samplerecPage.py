# -*- coding: utf-8 -*-
# @Time    : 2021/11/11
# @Author  : guanzhong.zhou
# @File    : 样本接收页面功能封装
from datetime import datetime

from openpyxl import load_workbook
from PageElemens.sampleRec_ele import *
from common.DataBaseConfig import executeSql
from common.editYaml import *
from common.all_path import pathologycheck_file_path, hstq_file_path, sampleprocessing_file_path, \
    wkdl_sr_file_path, sr_sample_imp_file, sr_sample_sublibrary_imp_file, orderNub_path, \
    app_a_file_path, SR_sample_for_import_path
from common.screenshot import Screenshot
from common.xlsx_excel import add_write_excel_xlsx
from conf.config import create_lab_excel, specimen_list
from data.sql_action.execute_sql_action import ybjs_sql, get_sr_sample_lims, set_sr_sample_id_external
from uitestframework.basepageTools import BasePage
from common.logs import log
from uitestframework.exceptionsTools import TimeOutError
import re


class SampleReceivePage(BasePage):
    """
    样本接收类页面基础方法
    """

    def search_order(self):
        """
        查询已存在订单号，对该订单号进行接样
        """
        order = read_yaml(orderNub_path)  # 获取订单号
        log.info('接样页面点击搜索按钮')

        self.click_by_js('css', search_btn)
        self.sleep(0.5)

        log.info('搜索框录入订单号：%s' % order['order_number'])
        self.input('css', order_numb, order['order_number'])
        self.sleep(1)
        self.clicks('css', search_confirm)
        self.sleep(0.5)

        # 这里调用自定义截图方法
        Screenshot(self.driver).get_img("样本接收列表，点击搜索按钮，根据lims号搜索","搜索出订单号")

        self.clicks('xpath', chioce_result)
        self.sleep(0.5)
        log.info('选中搜索出的订单，进入样本接收详情页面')
        self.clicks('css', edit_sample_order)
        # 存在进入接样明细缓冲过长，这里加个页面刷新
        try:
            self.wait_loading()
        except TimeOutError as error1:
            log.warning('进入接样详情页面超时，刷新页面。{}'.format(error1))
            self.refresh()
        except Exception as error2:
            log.warning('进入接样详情页面报错，刷新页面。{}'.format(error2))
            self.refresh()

    def add_sample_type(self):
        """
        样本接收详情页面，增加样本项目信息
        """

        pro_name = self.get_text('xpath', project_name_chioce)  # 查看项目号是否有值，若没有则重新选择
        if len(pro_name) == 0:
            log.info('选择样本所属项目信息')
            self.clicks('xpath', project_name_chioce)
            self.wait_loading()
            self.input('xpath', project_name_input, 'J022')
            self.sleep(0.5)
            self.clicks('xpath', project_search_button)
            self.wait_loading()
            self.clicks('xpath', chioce_project_result)
            self.wait_loading()
            self.wait_loading()
            self.click_by_js('xpath', back_button)
            self.wait_loading()
            self.sleep(1)

        testitem = self.get_text('xpath', test_item)  # 查看检测产品是否有值，若没有则重新选择
        log.info('选择样本所属检测产品')
        if len(testitem) == 0:
            self.clicks('xpath', test_item)
            self.wait_loading()
            self.input('xpath', product_input, '世和一号')
            self.clicks('xpath', product_search_btn)

            self.clicks('xpath', open_product)
            self.clicks('xpath', choice_product)
            self.sleep(0.5)
            if self.isElementExists('css', tipInfo):
                self.clicks('css', tipInfo)
            self.wait_loading()
            self.wait_loading()
            self.clicks('xpath', close_button)
            self.sleep(1)

    def chioce_specimen(self, specimenList):
        """
        封装选择样本类型方法
        """
        self.sleep(0.5)
        # 样本类型选择按钮，元素定位
        self.clicks('xpath', sample_TypeName)
        self.wait_loading()
        # 样本类型选项弹框，查询录入文本框
        self.input('xpath', original_specimen_type, specimenList)
        print('选择样本类型：', specimenList)
        self.sleep(0.5)
        self.clicks('xpath', search_type_button)
        log.info('输入样本类型，点击查询按钮，进行检索')
        self.wait_loading()
        self.clicks('xpath', chioce_search_result)
        self.sleep(0.5)
        self.clicks('xpath', specimen_type_comfirm)
        self.sleep(0.5)

    def sampleRec_sampleDetail(self):
        """
        新增五种样本并选择样本类型：
        FF白片
        ED抗凝血
        骨冷冻组织
        DNA文库
        外部血浆
        cfDNA文库
        """

        def set_sampleType():
            """设置样本类型方法封装"""
            for index in range(tips + 1, (tips + num) + 1):
                self.click_by_js('css', one_by_one_samples.format(index))  # 先选中对应样本
                self.sleep(0.5)
            self.chioce_specimen(s_type)
            for b in range(tips + 1, (tips + num) + 1):
                self.click_by_js('css', one_by_one_samples.format(b))
            self.sleep(1)

        log.info('添加5种样本，')
        total = specimen_list.values()
        n = 0
        for d in total:
            n += d
        # 新增样本
        s = 0
        while s < n:
            self.clicks('xpath', add_sample)
            print('新增第', s + 1, "条样本")
            self.wait_loading()
            s += 1
        lists = self.findelements('xpath', all_samples)
        print("统计新增样本总数：", len(lists))
        # 添加样本类型
        tips = 0
        for s_type, num in specimen_list.items():  # 从字典取值，样本类型以及对应的要添加的数量
            if s_type == 'FFPE白片':  # 跟据取值依次判断
                for i in range(tips, tips + num):  # 样本类型数量，循环选中多少数量的样本
                    j = i + 1
                    self.clicks('css', one_by_one_samples.format(j))  # 先选中对应样本
                self.chioce_specimen(s_type)  # 执行添加样本类型方法
                self.sleep(1)
                self.moved_to_element('css', add_pathology_worksheet)  # 移动鼠标至病理任务，添加HE任务
                self.sleep(1.5)
                self.clicks('xpath', pathology_worksheet_HE)
                s = self.get_text('xpath', pathology_worksheet_HE)

                print(s)
                self.wait_loading()
                self.sleep(1)
                for i in range(tips, tips + num):
                    j = i + 1
                    self.clicks('css', one_by_one_samples.format(j))
                    self.sleep(0.2)
                self.moved_to_element('css', add_pathology_worksheet)  # 移动鼠标至病理任务，添加HE任务
                self.sleep(1)
                self.click_by_js('xpath', pathology_worksheet_PD)
                s = self.get_text('xpath', pathology_worksheet_PD)
                print(s)
                self.wait_loading()
                self.sleep(1)
                print('添加病理成功')

            elif s_type == 'EDTA抗凝血':
                set_sampleType()

            elif s_type == '骨冷冻组织':
                set_sampleType()

            elif s_type == 'DNA文库':
                set_sampleType()

            elif s_type == '外部血浆':
                set_sampleType()

            elif s_type == 'cfDNA文库':
                set_sampleType()
            tips += num  # 每次循环取不同样本，所以各样本数量相加，取其排序下标，在页面中根据对应下标进行定位

        self.sleep(1)
        self.clicks('css', save_btn)
        self.wait_loading()

        # 这里调用自定义截图方法
        Screenshot(self.driver).get_img("样本接收详情页，点击新增样本，点击保存新增样本","保存新增样本成功")

    def generate_laboratory_process(self):
        """
        样本接收-样本明细(未审核),选择样本并生成实验流程，对应的样本与实验流程关系.实验流程需要和上一步样本类型一致
        FF白片——核酸提取-不破碎；
        ED抗凝血-体液样本分离；
        骨冷冻组织-21基因；
        DNA文库-文库定量；
        """

        def expProcess_planne(sampleTotal, sampleType, expTemp, expType):
            """把生成实验流程方法封装"""
            for ic in range(1, len(sampleTotal) + 1):  # 从所有存在于实验流程弹框中的样本，根据样本类型值，与列表取值进行比对
                specimenType = self.get_text('xpath', template_sample_type.format(ic))
                if specimenType == sampleType:  # 如果页面列表中，样本类型值与列表循环出的一致，则在页面弹框选中
                    self.clicks('xpath', one_by_one_chioce_sample.format(ic))  # 先选中对应样本
            self.clicks('xpath', laboratory_process_temp_btn)  # 点击选择实验流程按钮
            self.sleep(1)
            self.click_by_js('css', expTemp)  # 切换到Illumina或华大模版
            self.sleep(0.5)
            # 选中样本类型对应的实验流程
            self.clicks('xpath', LibProcessVisible.format(expType))
            print("选中")
            self.sleep(0.5)
            self.clicks('xpath', LibProcessVisible_btn)  # 选择实验流程弹框确认按钮
            self.sleep(0.5)
            self.clicks('xpath', laboratory_process_planned_btn)  # 选择探针
            if re.search(r'当前模板不需预设探针', self.get_source):
                for ib in range(1, len(sampleTotal) + 1):
                    specimenType = self.get_text('xpath', template_sample_type.format(ib))
                    if specimenType == sampleType:
                        self.clicks('xpath', one_by_one_chioce_sample.format(ib))
            else:
                self.sleep(0.5)
                self.clicks('xpath', laboratory_process_planned_chioce)
                self.sleep(0.5)
                self.clicks('xpath', laboratory_process_planned_comfirm)
                self.sleep(0.5)
                # 把前面已完成操作的样本进行取消选中，接下来选中其它类型的样本
                for ib in range(1, len(sampleTotal) + 1):
                    specimenType = self.get_text('xpath', template_sample_type.format(ib))
                    if specimenType == sampleType:
                        self.clicks('xpath', one_by_one_chioce_sample.format(ib))
            self.sleep(0.5)

        # laboratory_pro = ['核酸提取-破碎', '样本处理-样本分离', '21基因', '文库定量']
        log.info('选中样本生成实验流程')
        self.clicks('css', all_chioce)  # 全选样本
        self.sleep(1)
        self.clicks('xpath', generateLibProcessVisible)  # 点击生成实验流程
        self.wait_loading()
        self.sleep(0.5)

        # 这里调用自定义截图方法
        Screenshot(self.driver).get_img("样本接收，录入样本类型后，勾选样本，点击生成实验流程按钮","打开实验流程配置弹框")

        lists = self.findelements('xpath', template_sample_all)  # 实验流程弹框中，全部样本数量

        if lists:
            for s_type, num in specimen_list.items():  # 取出样本类型及其数量
                if s_type == 'FFPE白片':  # 判断从列表循环取出本的类型
                    expProcess_planne(lists, 'FFPE白片', Illumina, '核酸提取-破碎')

                elif s_type == 'EDTA抗凝血':
                    expProcess_planne(lists, 'EDTA抗凝血', Illumina, '样本处理-样本分离')

                elif s_type == '骨冷冻组织':
                    expProcess_planne(lists, '骨冷冻组织', Illumina, '21基因')

                elif s_type == 'DNA文库':#SR样本
                    expProcess_planne(lists, 'DNA文库', Illumina, '文库定量')

                elif s_type == '外部血浆':
                    expProcess_planne(lists, '外部血浆', Illumina, '提取-质谱仪上机')

                elif s_type == 'cfDNA文库':#SR样本
                    expProcess_planne(lists, 'cfDNA文库', huada, '华大-APP-A')

        # 获取当前窗口句柄
        now_handle = self.get_current_window_handle()
        print('获取当前窗口句柄', now_handle)

        # 点击生成实验流程，等待打开新的实验流程页面
        self.clicks('xpath', generatelaboratoryprocess_btn)
        self.wait_loading()
        self.sleep(2)
        # 获取所有窗口句柄
        all_handles = self.get_all_windows()
        for handle in all_handles:
            if handle != now_handle:
                # 切换到新窗口句柄，即新打开的流转表页面
                self.switch_to_window(handle)
                self.sleep(2)
                # 关闭新窗口句柄，关闭流转表页面
                self.close()
        self.sleep(1.5)
        # 切换到旧窗口句柄，回到原页面
        self.switch_to_window(now_handle)
        self.sleep(1)

    def save_all_samples_excel(self):
        """
        把所有样本信息保存到Excel中
        """
        log.info('样本信息保存到对应流程的Excel中')
        create_lab_excel()  # 初始化样本存储Excel文件
        # 准备sr样本数据
        self.sr_sample_import()
        lists = self.findelements('xpath', all_samples)
        self.sleep(0.5)

        def write_excel(filePath):
            """写入Excel方法封装"""
            data = []
            sample_lims = self.get_text('xpath', one_lims_num.format(i))  # lism号
            sample_lab = self.get_text('xpath', one_laboratory_num.format(i))  # 实验室号
            data.append(sample_lims)
            data.append(sample_lab)
            lims_list_values.append(data)
            add_write_excel_xlsx(filePath, lims_list_values)

        for i in range(1, len(lists) + 1):
            lims_list_values = []
            samples_type = self.get_text('xpath', one_sample_type.format(i))
            self.sleep(0.5)
            if samples_type == "FFPE白片":
                write_excel(hstq_file_path)
                add_write_excel_xlsx(pathologycheck_file_path, lims_list_values)
            elif samples_type == "EDTA抗凝血":
                write_excel(sampleprocessing_file_path)
            elif samples_type == "骨冷冻组织":
                write_excel(hstq_file_path)
            elif samples_type == "DNA文库":
                write_excel(wkdl_sr_file_path)
            elif samples_type == "外部血浆":
                write_excel(hstq_file_path)
            elif samples_type == "cfDNA文库":
                write_excel(app_a_file_path)

    def input_sampleamt(self):
        """
        录入样本包装量和样本计量
        """
        log.info('录入样本包装量')

        self.clicks('css', all_chioce)
        self.sleep(0.5)
        self.clicks('xpath', sample_PkgAmt)
        self.input('xpath', sample_PkgAmt_input, 1)
        self.clicks('xpath', sample_PkgAmt_input_comfirm_btn)
        self.sleep(0.5)

        log.info('录入样本计量')
        self.clicks('xpath', sample_amt)
        self.input('xpath', sample_amt_input, 20)
        self.clicks('xpath', sample_amt_input_comfirm_btn)
        self.sleep(0.5)



        log.info('录入样本基本数据后，样本信息保存')
        self.clicks('css', save_btn)
        self.wait_loading()
        self.refresh()

    def sr_sample_import(self):
        """
        准备sr样本数据，在sr信息登记模块使用。选取一条sr样本，设置sr样本的外部样本编号，并存入对应的导入模板
        """
        order = read_yaml(orderNub_path)  # 获取订单号
        log.info('录入样本备注')
        self.updata_sql(ybjs_sql.format(order['order_number']))
        log.info('获取订单接样表中所有的SR样本')
        sr_sample = executeSql.test_select_limsdb(get_sr_sample_lims.format(order['order_number']))
        sr_sampleLims = [list(i.values()) for i in sr_sample]
        print('接样的SR样本：', sr_sampleLims)

        # SR样本信息导入数据
        sr_sample_imp_data = [['J022'], ['靶向富集'], ['HiseqX'], ['PE150'], ['单'], ['不混样包Lane'], [1], ['G'],
                              ['单梯度绝对'],
                              [2100], [0.796],
                              [2.3], [24], [330], ['双标签'], ['是'], ['是'], [2], ['否'], [12], ['暂无'], [2], [34],
                              ['包埋文库'],
                              ['F类'],
                              ['备注'], [datetime.now().strftime('%Y.%m.%d')], [1], [2], ['否']]
        sr_sample_sublibrary_imp_data = [['O01-004'], ['DC-013'], ['AGTCT'], ['DC-014'], ['DC-015'], ['J022']]

        sr_sample_imp_file_r = load_workbook(sr_sample_imp_file)
        sr_sample_sublibrary_imp_file_r = load_workbook(sr_sample_sublibrary_imp_file)

        lims = []
        for val in range(len(sr_sampleLims)):

            sample_id_external = [sr_sampleLims[val][0] + '_TEST_SR']
            lims.append(sr_sampleLims[val][0])
            sr_sample_imp_data.insert(1, sample_id_external)
            sr_sample_sublibrary_imp_data.insert(0, sample_id_external)

            table = sr_sample_imp_file_r.active
            table1 = sr_sample_sublibrary_imp_file_r.active
            nrows = table.max_row  # 获得行数
            nrows1 = table1.max_row  # 获得行数
            # 注意行业列下标是从1开始的
            for i in range(1, len(sr_sample_imp_data) + 1):
                for j in range(1, len(sr_sample_imp_data[i - 1]) + 1):
                    table.cell(nrows + 1, i).value = sr_sample_imp_data[i - 1][j - 1]
            sr_sample_imp_data.pop(1)

            for i in range(1, len(sr_sample_sublibrary_imp_data) + 1):
                for j in range(1, len(sr_sample_sublibrary_imp_data[i - 1]) + 1):
                    table1.cell(nrows1 + 1, i).value = sr_sample_sublibrary_imp_data[i - 1][j - 1]
            sr_sample_sublibrary_imp_data.pop(0)
            executeSql.test_updateByParam(
                set_sr_sample_id_external.format(sample_id_external[0], sr_sampleLims[val][0]))
        sr_sample_imp_file_r.save(sr_sample_imp_file)
        sr_sample_sublibrary_imp_file_r.save(sr_sample_sublibrary_imp_file)

        # 将sr样本保存到yaml文件，在SR样本信息登记模块使用
        datas = read_yaml(SR_sample_for_import_path)
        datas["rec_sr_sample_for_sr_import"] = lims
        save_yaml(SR_sample_for_import_path, datas)

    def submit_sample_for_review(self):
        """
        对样本进行批量提交审核
        """
        log.info('选中样本，点击批量提交审核')
        self.clicks('css', all_chioce)
        self.sleep(0.5)
        self.clicks('css', batch_submit_for_review)
        self.wait_loading()
        self.sleep(2)

    def get_save_info(self):
        """
        获取数据操作后，页面给出的提示信息语
        Submit successfully
        review successfully
        """
        log.info('获取页面提示信息')
        if self.isElementExists('css', save_info):
            return self.get_text('css', save_info)
        else:
            return None

    # 切换登录用户进行审核
    def batch_submit_for_review(self):
        """
        切换登录用户进行审核
        """
        log.info('点击页面退出登录按钮，切换登录用户')
        self.click_by_js('css', logout_btn)
        self.sleep(1)
        self.click_by_js('xpath', logout_choice)
        self.sleep(1)

    def batch_review(self):
        """
        切换待审核页面，查看待审核样本，并进行样本审核
        """
        log.info('登录后，切换到样本待审核页面')
        self.clicks('xpath', review_pending)
        self.wait_loading()
        self.sleep(1)
        log.info('待审核页面选中待审核样本，并点击批量审核按钮')
        # 全选样本数据
        self.clicks('xpath', review_all_chioce)
        self.sleep(1)
        # 点击批量完成审核按钮
        self.clicks('xpath', batch_checked_for_review)
        self.sleep(0.5)

        # 这里调用自定义截图方法
        Screenshot(self.driver).get_img("样本接收，勾选待审核样本，点击批量完成审核，录入审核信息后点击确认","审核样本成功")

        # 弹出框录入用户密码
        self.input('xpath', password_inpt, 1)
        # 点击下一步
        self.clicks('xpath', next_step)
        # 弹出框录入审核理由
        self.input('xpath', review_remarks, '测试')
        # 点击提交
        self.clicks('xpath', review_confirm)

        self.wait_loading()

