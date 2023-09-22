# -*- coding: utf-8 -*-
# @Time    : 2021/12/09
# @Author  : guanzhong.zhou
# @File    : 文库定量模块页面功能封装
from datetime import datetime
import pyperclip
import yaml, xlrd
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys
from common import editYaml
from common.all_path import wkdl_sr_file_path, wkdl_sequencing_group_number, wkdl_non_sr_file_path, \
    functionpageURL_path, wkdl_detail_sr_import_path, wkdl_detail_non_sr_import_path, \
    wkdl_detail_single_gradient_sr_import_path, wkdl_detail_single_gradient_non_sr_import_path, position_in_box_path, \
    wkdl_result_non_sr_import_path, wkdl_result_sr_import_path, wkdl_result_standard_sample_path, wkdl_hdsr_file_path
from common.screenshot import Screenshot
from common.DataBaseConfig import executeSql
from common.xlsx_excel import read_excel_col, pandas_write_excel, get_lims_for_excel
from PageElemens.libraryquantification_ele import *
from conf.config import libquant_result
from data.sql_action.execute_sql_action import wkdl_detail_sql1, wkdl_detail_sql2, wkdl_detail_sql3, wkdl_result_sql1, \
    dl_next_step, update_wkdl_result_hd_data, update_wkdl_result_data
from uitestframework.basepageTools import BasePage
from common.logs import log


class LibraryQuantificationPage(BasePage):
    """
    文库定量页面方法封装
    """

    # 打开指定页面
    def enter_function_page(self, url):
        """
        进入指定url页面
        """
        js = 'window.open("{}");'
        self.executeJscript(js.format(url))
        self.wait_loading()
        self.sleep(1)

    # 获取页面提示信息
    def get_pageinfo(self):
        """
        获取页面操作提示信息
        Task list saved successfully---保存样本到任务单成功
        Submit successfully---提交成功
        sample in storage successfully---入库成功
        """

        return self.get_text('xpath', page_success_info)

    # 新增任务单
    def add_task(self):
        """
        新建文库定量任务单
        """
        # 判断是否存在建库电子实验记录待签字交接提示信息
        # if self.isElementExists('css', electronic_experiment_record_signed_tips):
        #     self.clicks('css', electronic_experiment_record_signed_tips)
        log.info("文库定量首页，点击新建按钮，进入样本待选表，新增文库定量任务")
        self.clicks('css', add_sample_process_task)
        self.wait_loading()

        log.info("选择定量类型")
        self.clicks('css', quantification_type)
        self.sleep(1)
        self.clicks('xpath', quantification_type_choice)
        self.wait_loading()
        self.sleep(0.5)

        log.info(" 添加任务，描述")
        self.input('css', task_description, '文库定量自动化测试')

        log.info("选择sop")
        self.clicks('css', select_sop)
        self.sleep(0.5)
        self.clicks('css', select_sop_choice)
        self.wait_loading()

    # excel中获取非SR待选样本
    def get_non_sr_sample_from_excel(self, col_name):
        """
         从对应的Excel中获取上一步流传下来的本节点的待选表lims样本号
        :param col_name: excel列名
        :return:lims_nub
        """

        log.info("从上一步流转Excel中获取非SR数据")
        lims_nub = read_excel_col(wkdl_non_sr_file_path, col_name)
        if lims_nub:
            return lims_nub
        else:
            return None

    # excel中获取SR待选样本
    def get_sr_sample_from_excel(self):
        """
        获取接样Excel中的所有sr样本，从数据库根据lims号获取富集lims号
        :return:sr_lims_id_str列表
        """
        log.info(" 从接样表获取所有样本，到数据库筛选出SR样本的样本lims号")
        enrichment_nub = read_excel_col(wkdl_sr_file_path, 'lims号')
        if enrichment_nub:
            sample_datas = str(enrichment_nub).replace('[', '').replace(']', '')  # 把取出的列表转换成sql条件中的元组格式
            # 根据接样lims号，到数据库中（sample_info_t表）获取定量富集lims号
            sa = executeSql.test_select_limsdb(
                "SELECT DISTINCT (pooling_lims_id) FROM sample_info_t t1 WHERE t1.original_sample_id_lims IN ({})AND "
                "t1.sr_type = '02'".format(sample_datas))
            result_sql = [item[key] for item in sa for key in item]  # sql返回值是列表套字典，这里取出字典的值
            while None in result_sql:  # 取出字典值中的空值None
                result_sql.remove(None)
            print('样本接收定量流程SR样本', result_sql)
            return result_sql
        else:
            return None

    # excel中获取华大SR待选样本
    def get_huada_sr_sample(self):
        """获取华大流程的sr样本"""
        log.info("APP-A待选表核对lims号功能，并保存任务单号")
        lims_id_str = get_lims_for_excel(wkdl_hdsr_file_path)
        log.info('获取接样流程中的lims样本号,核对lims号是否存在')
        self.clicks('css', check_lims_sample_num)
        self.sleep(1)
        self.input('css', check_lims_sample_number_textarea, lims_id_str)
        self.sleep(0.5)
        self.clicks('css', check_lims_sample_number_confirm)
        self.wait_loading()
        info=self.choice_sample()
        return info

    # 待选表校验lims号
    def check_lims_num(self):
        """
        待选表核对lims号功能，并保存任务单号
        """
        log.info(" 获取接样流程中的sr样本,核对富集lims是否存在")
        non_sr_lims = self.get_non_sr_sample_from_excel('lims号')
        sr_lims = self.get_sr_sample_from_excel()
        all_lims_id = None
        if non_sr_lims is not None and sr_lims is not None:
            lims_id_str = "\n".join(non_sr_lims)  # 取出Excel表中样本，拼接成字符串录入到检索文本中
            sr_lims_id_str = "\n".join(sr_lims)  # 把所有值（富集lims号）转换为带换行的字符串

            all_lims_id = lims_id_str + "\n" + sr_lims_id_str
        elif non_sr_lims is not None:
            lims_id_str = "\n".join(non_sr_lims)
            all_lims_id = lims_id_str

        elif sr_lims is not None:
            sr_lims_id_str = "\n".join(sr_lims)
            all_lims_id = sr_lims_id_str
        else:
            print("Excel里没有数据啊，搞咩啊！！！")
        log.info('待选表核对样本')
        if all_lims_id:
            self.clicks('css', check_lims_sample_num)
            self.sleep(1)
            self.input('css', check_lims_sample_number_textarea, all_lims_id)
            self.sleep(0.5)
            self.clicks('css', check_lims_sample_number_confirm)
            self.wait_loading()
            info=self.choice_sample()
        else:
            raise Exception("没有取到数据，请检查！！")
        return info
    # 选中核对后的样本，点击【加入选中样本&保存】
    def choice_sample(self):
        # ('选中核对后的样本，点击【加入选中样本&保存】')
        self.click_by_js('css', all_choice)
        self.sleep(0.5)
        self.clicks('css', addSelect_or_save_btn)
        # 调用自定义截图方法
        Screenshot(self.driver).get_img("文库定量待选表点击核对lims号，录入样本号进行查询，勾选查询结果，并保存任务单号","保存任务单成功")
        pageinfo = self.get_pageinfo()
        self.wait_loading()
        # 判断子sop样本数据数否需要录入
        result = self.driver.execute_script(
            'return document.getElementsByClassName("commonTaskDetail-commonTaskDetailBtn-sopSampleNumber")[0].disabled')
        if not result:
            self.clicks('css', sopSampleNumber_btn)  # 点击子sop样本数量
            self.wait_loading()
            self.clicks('css', sopSampleNumber)  # 子sop样本数量弹框中点击数量录入框
            self.input('css', sopSampleNumber_input, 10)
            self.clicks('css', sopSampleNumber_confirm)
            self.wait_loading()
        return pageinfo

    # 进入明细表或结果表
    def enter_result_list(self, ele, page):
        """
        待选表或者明细表，在进入下一节点时，获取下一节点URL地址，并写入临时数据文件中
        :param ele: 点击进入下一节点元素定位
        :param page: 下一节点页面名称
        """
        urldata = editYaml.read_yaml(functionpageURL_path)

        self.clicks('css', ele)
        log.info('点击按钮进入{}'.format(page))
        self.wait_loading()
        self.wait_loading()
        self.sleep(2)

        url = self.get_current_url()  # 获取当前页面URL地址
        print('获取的URL地址', url)
        urldata["url"] = url
        with open(functionpageURL_path, 'w', encoding='utf-8') as fs:  # 写入模式获取的URL地址到yaml文件中
            yaml.safe_dump(urldata, fs, allow_unicode=True)
        print("写入后的URL地址", urldata)

    # 准备SR类型批量粘贴导入的数据
    def read_sr_import_data(self, sr_samples, import_file_path):
        """
        根据sr样本数量，修改sr类型导入模板
        """
        log.info(" 获取SR样本")

        wb = load_workbook(filename=import_file_path)  # 打开excel文件
        # 把流程中的样本号写入待导入模板
        for i in range(0, len(sr_samples)):
            k = i + 2
            self.sleep(0.5)
            ws = wb.active
            # 根据需要修改表格数据
            ws.cell(k, 1, sr_samples[i])  # 修改第k行，第index列值
        wb.save(import_file_path)
        self.sleep(1)
        # 从模板中复制出修改后的待导入数据
        data = xlrd.open_workbook(import_file_path)
        num_list = []
        for index in range(0, len(sr_samples)):
            h = index + 1
            tables = data.sheets()[0]
            # allrows = tables.nrows
            vals = tables.row_values(h)
            imp_data = '\t'.join(map(str, vals))
            num_list.append(imp_data)
        print("\n".join(map(str, num_list)))
        pyperclip.copy("\n".join(map(str, num_list)))

    # 准备非SR类型批量粘贴导入的数据
    def read_non_sr_import_data(self, non_sr_samples_name, import_file_path, enrichments_name, Sequencing_group_number):
        """
          根据非sr样本数量，修改非sr类型导入模板
        :param non_sr_samples_name: Excel读取非SR样本文库名称
        :param import_file_path: 导入模板路径
        :param enrichments_name:系统中，文库名称的定位
        :param Sequencing_group_number:系统中，上机分组号的定位
        """

        taskstatus = self.get_text('css', detail_task_id)  # 获取任务单号
        lims_id = executeSql.test_select_limsdb(wkdl_detail_sql1.format(taskstatus[5:].strip()))  # 从数据库获取当前任务单号下样本总数
        lims_count = [item[key] for item in lims_id for key in item]

        wb = load_workbook(filename=import_file_path)  # 打开excel文件
        ws = wb.active
        log.info("把流程中的非sr样本号写入待导入模板")
        sequencing_group_number_list = []

        row = 2
        for g in range(0, lims_count[0]):
            k = g + 1
            # 依次取出明细表中文库名称，与Excel流转数据中的文库名称进行比对，如果与流转表中的一致则为非sr样本
            lims_num = self.get_text('css', enrichments_name.format(k))
            if lims_num in non_sr_samples_name:
                self.sleep(0.5)
                sequencing_group_number = self.get_text('css', Sequencing_group_number.format(k))  # 获取上机分组号
                sequencing_group_number_list.append(sequencing_group_number)
                ws.cell(row, 2, sequencing_group_number)  # 修改第row行，第index列值,上机分组号
                ws.cell(row, 1, lims_num)  # 修改第row行，第index列值，富集lims号
                row += 1
        log.info(" 获取的非sr样本上机分组号写入到yaml文件中，在结果表中取用，用以区分sr与非sr")
        non_sr_group_number_list = list(set(sequencing_group_number_list))
        with open(wkdl_sequencing_group_number, 'w', encoding='utf-8') as fs:
            yaml.dump_all([non_sr_group_number_list], fs, allow_unicode=True)

        wb.save(import_file_path)
        self.sleep(1)
        # 从模板中复制出修改后的待导入数据
        data = xlrd.open_workbook(import_file_path)
        num_list = []
        for nss in range(0, len(non_sr_samples_name)):
            h = nss + 1
            tables = data.sheets()[0]
            vals = tables.row_values(h)
            imp_data = '\t'.join(map(str, vals))
            num_list.append(imp_data)
        print("\n".join(map(str, num_list)))
        pyperclip.copy("\n".join(map(str, num_list)))

    # 明细表入库信息、批量粘贴导入实验数据
    def detail_libraryenrichment(self):
        """
        文库定量明细表，明细表批量入库类型、sr/非sr样本批量数据粘贴导入
        """

        log.info("明细表选择样本入库类型")
        self.clicks('css', detail_all_choice)  # 列表全选按钮
        self.sleep(0.5)


        log.info(" 文库定量明细表批量粘贴导入SR数据")
        sr_samples = self.get_sr_sample_from_excel()
        non_sr_samples = self.get_non_sr_sample_from_excel('文库名称')  # 获取非SR样本文库名称
        if sr_samples is not None:
            self.read_sr_import_data(sr_samples, wkdl_detail_sr_import_path)
            self.clicks('css', sr_data_batch_paste_import)  # 明细表SR数据批量粘贴导入按钮
            self.sleep(0.2)
            self.clicks('css', sr_data_batch_paste_import_input)
            self.findelement('css', sr_data_batch_paste_import_input).send_keys(Keys.CONTROL, 'v')
            self.sleep(0.2)
            self.clicks('css', sr_data_batch_paste_import_comfirm)  # 明细表批量粘贴导入弹框确认按钮
            self.sleep(1)
        log.info("文库定量明细表批量粘贴导入非SR数据")
        if non_sr_samples is not None:
            self.read_non_sr_import_data(non_sr_samples, wkdl_detail_non_sr_import_path, enrichment_name,
                                         all_Sequencing_group_number)
            self.click_by_js('css', non_sr_data_batch_paste_import)  # 明细表SR数据批量粘贴导入按钮
            self.sleep(0.2)
            self.clicks('css', non_sr_data_batch_paste_import_input)
            self.findelement('css', non_sr_data_batch_paste_import_input).send_keys(Keys.CONTROL, 'v')
            self.sleep(0.2)
            self.clicks('css', non_sr_data_batch_paste_import_comfirm)  # 明细表批量粘贴导入弹框确认按钮
            self.sleep(0.5)

        log.info(" 导入成功后点击批量计算")
        self.clicks('css', auto_calculate)
        self.sleep(1)

    # 明细表导入华大样本数据
    def detail_import_huada_sr_sample(self):
        """导入华大样本数据"""
        sr_samples = read_excel_col(wkdl_hdsr_file_path, 'lims号')

        self.read_sr_import_data(sr_samples, wkdl_detail_sr_import_path)
        self.clicks('css', sr_data_batch_paste_import)  # 明细表SR数据批量粘贴导入按钮
        self.sleep(0.2)
        self.clicks('css', sr_data_batch_paste_import_input)
        self.findelement('css', sr_data_batch_paste_import_input).send_keys(Keys.CONTROL, 'v')
        self.sleep(0.2)
        self.clicks('css', sr_data_batch_paste_import_comfirm)  # 明细表批量粘贴导入弹框确认按钮
        self.sleep(1)
        self.clicks('css', detail_all_choice)
        self.sleep(0.5)
        log.info(" 导入成功后点击批量计算")
        self.clicks('css', auto_calculate)
        self.sleep(1)

    # 明细表sr数据选择测序仪、预计上机时间、FC代码、预设Lane号
    def sequencer_estimatedsequencingdate_fccode_preset_Lanenumber(self):
        """
         明细表sr数据选择测序仪、预计上机时间、FC代码、预设Lane号
        """
        sr_lims = self.get_sr_sample_from_excel()  # 调用方法读取所有SR样本lims号
        if sr_lims is not None:
            log.info("在批量选中样本弹框中，录入sr样本号，筛选出样本，批量录入测序仪、预计上机时间、FC代码、预设Lane号")
            sr_lims_id_str = "\n".join(sr_lims)  # 把所有值（富集lims号）转换为带换行的字符串
            self.clicks('css', batch_select_sample)
            self.sleep(0.5)
            log.info(" 批量选择样本弹框录入文本框")
            self.input('css', batch_select_sample_input, sr_lims_id_str)
            self.clicks('css', batch_select_sample_comfirm)
            self.detail_batch_data()

    # 明细表sr数据选择测序仪、预计上机时间、FC代码、预设Lane号
    def detail_batch_data(self):
        """选择测序仪、预计上机时间、FC代码、预设Lane号"""
        str_time = datetime.now().strftime('%Y.%m.%d')  # 获取当前时间
        log.info(" 选择测序仪")
        self.clicks('css', select_sequencer)
        self.wait_loading()
        self.clicks('css', select_sequencer_choice)
        self.clicks('css', select_sequencer_comfirm)
        self.sleep(1)
        log.info("选择FC代码")
        self.moved_to_element('css', select_fc_code)
        self.sleep(1)
        self.clicks('xpath', select_fc_code_choice)
        log.info(" 填入上机信息")
        self.clicks('css', sequencing_information)
        self.input('css', estimated_sequencing_date, str_time)
        self.wait_loading()
        self.input('css', preset_lane_number, 1)
        self.clicks('css', sequencing_information_comfirm)
        self.wait_pageinfo_end()
        self.sleep(0.5)
        log.info(" 点击生成上机分组号")
        self.clicks('css', create_sequencing_group_number)
        self.wait_loading()

    # 明细表华大样本生成生成结果
    def detail_huada_sr_sample_generate_product(self):
        """华大样本生成生成结果"""
        log.info(" 点击生成生成结果")
        # self.clicks('css', detail_all_choice)  # 全选按钮
        self.clicks('css', create_result)
        self.wait_loading()

    # 明细表生成结果操作
    def create_result_and_sequencing_group_number(self):
        """
        明细表对已录入测序仪、预计上机时间、FC代码、预设Lane号的sr样本，及非sr样本，生成结果操作
        """
        enrichment_nub = read_excel_col(wkdl_sr_file_path, 'lims号')  # 调用方法读取所有SR样本lims号
        lims_nub = read_excel_col(wkdl_non_sr_file_path, "文库名称")

        if enrichment_nub and lims_nub:
            self.clicks('css', detail_all_choice2)  # 全选按钮
            self.clicks('css', create_result)
            self.wait_loading()
        else:
            self.clicks('css', create_result)

            # 调用自定义截图方法
            Screenshot(self.driver).get_img("文库定量明细表点击生成结果按钮","生成结果成功")

            self.wait_loading()

    # 明细表单梯度定量，华大SR数据导入
    def hd_single_gradient_quantification(self):
        """明细表单梯度定量，华大SR数据导入"""

        log.info(" 华大单梯度定量SR数据导入")
        sr_samples = read_excel_col(wkdl_hdsr_file_path, 'lims号')

        if sr_samples is not None:
            self.read_sr_import_data(sr_samples, wkdl_detail_single_gradient_sr_import_path)  # 调用写入方法
            self.clicks('css', single_gradientsr_data_batch_paste_import)  # 点击单梯度导入sr按钮
            self.clicks('css', single_gradientsr_sr_data_batch_paste_import_input)
            self.findelement('css', single_gradientsr_sr_data_batch_paste_import_input).send_keys(Keys.CONTROL, 'v')
            self.sleep(0.5)
            self.clicks('css', single_gradientsr_sr_data_batch_paste_import_comfirm)  # 明细表批量粘贴导入弹框确认按钮
            self.sleep(1)

    # 明细表单梯度定量，SR、非SR数据导入
    def single_gradient_quantification(self):
        """明细表单梯度定量，SR、非SR数据导入"""
        log.info(" 单梯度SR数据导入")
        sr_samples = self.get_sr_sample_from_excel()  # 获取SR样本
        non_sr_samples = self.get_non_sr_sample_from_excel('文库名称')  # 获取非SR样本文库名称
        if sr_samples is not None:
            self.read_sr_import_data(sr_samples, wkdl_detail_single_gradient_sr_import_path)  # 调用写入方法
            self.clicks('css', single_gradientsr_data_batch_paste_import)  # 点击单梯度导入sr按钮
            self.clicks('css', single_gradientsr_sr_data_batch_paste_import_input)
            self.findelement('css', single_gradientsr_sr_data_batch_paste_import_input).send_keys(Keys.CONTROL, 'v')
            self.sleep(0.5)
            self.clicks('css', single_gradientsr_sr_data_batch_paste_import_comfirm)  # 明细表批量粘贴导入弹框确认按钮
            self.sleep(1)

        log.info(" 单梯度非SR数据导入")
        if non_sr_samples is not None:
            self.read_non_sr_import_data(non_sr_samples, wkdl_detail_single_gradient_non_sr_import_path,
                                         enrichment_name, all_Sequencing_group_number)  # 调用写入方法
            self.clicks('css', single_gradientsr_non_sr_data_batch_paste_import)  # 点击单梯度导入非sr按钮
            self.clicks('css', single_gradientsr_non_sr_data_batch_paste_import_input)
            self.findelement('css', single_gradientsr_non_sr_data_batch_paste_import_input).send_keys(Keys.CONTROL, 'v')
            self.sleep(0.5)
            self.clicks('css', single_gradientsr_non_sr_data_batch_paste_import_comfirm)  # 明细表批量粘贴导入弹框确认按钮
            self.sleep(1)

        log.info(" 保存表单数据")
        self.clicks('css', detail_save_result)
        self.wait_loading()
        self.sleep(1)

        log.info(" 调用数据库，执行写入余样包装量")
        task_id = self.get_text('css', detail_task_id)
        print(task_id)
        self.updata_sql(wkdl_detail_sql2.format(task_id[5:].strip()))
        # 刷新页面
        self.refresh()

    # 明细表提交
    def detail_sumbit(self):
        """
        文库定量明细表，样本提交操作
        :return:
        """
        self.clicks('css', detail_all_choice)  # 列表全选按钮
        self.sleep(0.5)
        self.clicks('css', detail_submit)  # 提交按钮
        self.sleep(0.5)
        # 调用自定义截图方法
        Screenshot(self.driver).get_img("文库定量明细表点击提交按钮","弹出提交确认按钮")

        self.clicks('css', detail_submit_comfirm)  # 提交弹框确认按钮
        self.wait_loading()
        self.sleep(1)

    # 明细表入库
    def detail_into_storage(self):
        """
        文库定量明细表样本入库操作
        """
        log.info("文库定量明细表，样本入库操作")
        self.clicks('css', detail_all_choice)  # 列表全选按钮
        self.sleep(0.5)
        self.clicks('css', deposit_into_storage)  # 入库按钮
        self.sleep(1)
        self.clicks('css', storage_all_choice)  # 入库弹框全选按钮


        log.info("文库定量明细表，样本入库选择入样本盒")
        self.clicks('css', batch_paste_sample_box)  # 入库弹框选择样本盒按钮
        self.wait_loading()
        self.input('css', target_storage, '自动化测试用(勿删)')
        self.sleep(0.5)
        self.clicks('css', select_sample_box_search)
        self.wait_loading()
        self.clicks('css', select_sample_box_choice)  # 入库弹框选选择样本盒值，默认选择列表第一条数据
        self.clicks('css', select_sample_box_comfirm)  # 入库弹框选选择样本盒弹框，确认按钮
        self.sleep(1)

        log.info("文库定量明细表，样本入库录入盒内位置")

        taskstatus = self.get_text('css', detail_task_id)  # 获取任务单号
        lims_id = executeSql.test_select_limsdb(wkdl_detail_sql3.format(taskstatus[5:].strip()))  # 从数据库获取当前任务单号下样本lims号

        lims_list = [item[key] for item in lims_id for key in item]  # 把获取的lims号转换为一维列表
        nub_list = [str(i) for i in range(1, len(lims_list) + 1)]  # 根据lims样本数量，生成数字列表，作为盒内位置编号用
        res = [list(i) for i in zip(lims_list, nub_list)]  # 将lims号和数字编号转换为二维列表格式，写入Excel
        print(res)
        pandas_write_excel(res, position_in_box_path)  # 把样本号和盒内位置编号写入Excel模板

        data = xlrd.open_workbook(position_in_box_path)  # 从Excel读取模板样本号和盒内位置编号
        num_list = []
        for index in range(0, len(lims_list)):
            tables = data.sheets()[0]
            vals = tables.row_values(index)
            imp_data = '\t'.join(map(str, vals))
            num_list.append(imp_data)
        print("\n".join(map(str, num_list)))
        pyperclip.copy("\n".join(map(str, num_list)))

        log.info("粘贴到【批量粘贴盒内位置】文本框")
        self.clicks('css', batch_copy_BoxPosition)
        self.findelement('css', batch_copy_BoxPosition_input).send_keys(Keys.CONTROL, 'v')
        self.sleep(0.5)
        self.clicks('css', batch_copy_BoxPosition_comfirm)
        self.sleep(1)
        # 调用自定义截图方法
        Screenshot(self.driver).get_img("文库定量明细表点击入库按钮，在弹框中录入库位信息和盒内位置后点击下一步","样本入库成功")

        self.clicks('css', storage_next)
        self.wait_loading()

        self.executeJscript('document.getElementsByClassName("vxe-table--body-wrapper")[0].scrollLeft=3200')
        self.sleep(0.5)
        pageinfo = self.get_text('css', detail_sumbit_status)
        print(pageinfo)
        return pageinfo

    # 结果表华大样本导入数据
    def result_huada_sr_sample_write_import_file(self):
        """结果表华大样本写入导入模版"""
        taskstatus = self.get_text('css', result_task_id)  # 获取任务单号
        executeSql.test_updateByParam(update_wkdl_result_hd_data.format(taskstatus[5:].strip()))
        self.refresh()
        wb_sr = load_workbook(filename=wkdl_result_sr_import_path)  # 打开SR文件
        lims_id = executeSql.test_select_limsdb(wkdl_result_sql1.format(taskstatus[5:].strip()))  # 从数据库获取当前任务单号下样本总数
        lims_count = [item[key] for item in lims_id for key in item]
        ws = wb_sr.active
        sr_row = 2
        for sp in range(lims_count[0]):  # 从数据库获取当前任务单号下样本总数
            lims_id = self.get_text('css', result_samples__number.format(sp + 1))  # 获取页面lims号
            ws.cell(sr_row, 1, lims_id)  # 修改第row行，第index列值，富集lims号
            sr_row += 1
        wb_sr.save(wkdl_result_sr_import_path)
        self.copy_import_data(lims_count[0], wkdl_result_sr_import_path)
        self.clicks('css', result_sr_data_batch_paste_import)
        self.sleep(0.5)
        self.findelement('css', result_sr_data_batch_paste_import_input).send_keys(Keys.CONTROL, 'v')
        self.sleep(1)
        self.clicks('css', result_sr_data_batch_paste_import_comfirm)
        self.sleep(1)
        self.clicks('css', result_save)
        self.wait_loading()
    # 结果表分别根据sr样本lims号、非sr样本分组号，写入模板
    def result_data_write_import_file(self):
        """
        文库定量结果表导入粘贴非SR数据
        """
        taskstatus = self.get_text('css', result_task_id)  # 获取任务单号
        executeSql.test_updateByParam(update_wkdl_result_data.format(taskstatus[5:].strip()))
        self.refresh()

        wb_non_sr = load_workbook(filename=wkdl_result_non_sr_import_path)  # 打开非SR数据文件
        wb_sr = load_workbook(filename=wkdl_result_sr_import_path)  # 打开SR文件
        ws = wb_sr.active
        wns = wb_non_sr.active

        lims_id = executeSql.test_select_limsdb(wkdl_result_sql1.format(taskstatus[5:].strip()))  # 从数据库获取当前任务单号下样本总数
        lims_count = [item[key] for item in lims_id for key in item]

        log.info("读出存在yaml文件中的临时非SR样本上机分组号")
        with open(wkdl_sequencing_group_number, 'r+', encoding='utf-8') as f:
            cfg = f.read()
            non_srgroup_nums = yaml.load_all(cfg, Loader=yaml.FullLoader)
        non_srgroup_num_list = [key for i in non_srgroup_nums for key in i]

        non_row = 2
        sr_row = 2

        for sp in range(lims_count[0]):  # 从数据库获取当前任务单号下样本总数
            sap_gop_num = self.get_text('css', result_samples_group_number.format(sp + 1))  # 获取页面上机分组号
            print(sap_gop_num)
            lims_id = self.get_text('css', result_samples__number.format(sp + 1))  # 获取页面lims号
            print(lims_id)
            log.info(" 取出所有样本的分组号，判断是否包含包存在yaml中的非SR样本分组号，若包含则把分组号写入非sr模板，不包含则取出该分组号对应的lims号写入sr模板")
            if sap_gop_num in non_srgroup_num_list:
                wns.cell(non_row, 1, sap_gop_num)  # 修改第row行，第index列值,上机分组号
                non_row += 1
            else:
                ws.cell(sr_row, 1, lims_id)  # 修改第row行，第index列值，富集lims号
                sr_row += 1
        # 保存模板文件
        wb_non_sr.save(wkdl_result_non_sr_import_path)
        wb_sr.save(wkdl_result_sr_import_path)

    # 复制模板数据
    def copy_import_data(self, length, import_file):
        """
        复制模板数据方法封装
        :param length: 复制模板文件的行数
        :param import_file: 接收一个打开的模板文件变量
        """
        data = xlrd.open_workbook(import_file)
        num_list = []
        for nss in range(0, length):
            h = nss + 1
            tables = data.sheets()[0]
            # allrows = tables.nrows
            vals = tables.row_values(h)
            imp_data = '\t'.join(map(str, vals))
            num_list.append(imp_data)
        print("\n".join(map(str, num_list)))
        pyperclip.copy("\n".join(map(str, num_list)))

    # 从模板取出数据，批量粘贴导入
    def non_sr_and_sr_data_paste_import(self):
        """
        从模板取出数据，批量粘贴导入
        """
        taskstatus = self.get_text('css', result_task_id)  # 获取任务单号
        lims_id = executeSql.test_select_limsdb(wkdl_result_sql1.format(taskstatus[5:].strip()))  # 从数据库获取当前任务单号下样本总数
        lims_count = [item[key] for item in lims_id for key in item]

        # 读出存在yaml文件中的临时非SR样本上机分组号
        lims_nub = read_excel_col(wkdl_non_sr_file_path, "文库名称")
        enrichment_nub = read_excel_col(wkdl_sr_file_path, 'lims号')
        if lims_nub and enrichment_nub:
            with open(wkdl_sequencing_group_number, 'r+', encoding='utf-8') as f:
                cfg = f.read()
                non_srgroup_nums = yaml.load_all(cfg, Loader=yaml.FullLoader)
            non_srgroup_num_list = [key for i in non_srgroup_nums for key in i]

            # 批量粘贴非SR样本
            length1 = len(non_srgroup_num_list)
            self.copy_import_data(length1, wkdl_result_non_sr_import_path)
            self.clicks('css', result_non_sr_data_batch_paste_import)
            self.sleep(0.5)
            self.findelement('css', result_non_sr_data_batch_paste_import_input).send_keys(Keys.CONTROL, 'v')
            self.sleep(1)
            self.clicks('css', result_non_sr_data_batch_paste_import_comfirm)
            self.sleep(1)

            # 批量粘贴SR样本
            length2 = lims_count[0] - len(non_srgroup_num_list)
            self.copy_import_data(length2, wkdl_result_sr_import_path)
            self.clicks('css', result_sr_data_batch_paste_import)
            self.sleep(0.5)
            self.findelement('css', result_sr_data_batch_paste_import_input).send_keys(Keys.CONTROL, 'v')
            self.sleep(1)
            self.clicks('css', result_sr_data_batch_paste_import_comfirm)
            self.sleep(1)

        elif lims_nub:
            with open(wkdl_sequencing_group_number, 'r+', encoding='utf-8') as f:
                cfg = f.read()
                non_srgroup_nums = yaml.load_all(cfg, Loader=yaml.FullLoader)
            non_srgroup_num_list = [key for i in non_srgroup_nums for key in i]

            # 批量粘贴非SR样本
            length1 = len(non_srgroup_num_list)
            self.copy_import_data(length1, wkdl_result_non_sr_import_path)
            self.clicks('css', result_non_sr_data_batch_paste_import)
            self.sleep(0.5)
            self.findelement('css', result_non_sr_data_batch_paste_import_input).send_keys(Keys.CONTROL, 'v')
            self.sleep(1)
            self.clicks('css', result_non_sr_data_batch_paste_import_comfirm)
            self.sleep(1)
        elif enrichment_nub:

            # 批量粘贴SR样本
            self.copy_import_data(lims_count[0], wkdl_result_sr_import_path)
            self.clicks('css', result_sr_data_batch_paste_import)
            self.sleep(0.5)
            self.findelement('css', result_sr_data_batch_paste_import_input).send_keys(Keys.CONTROL, 'v')
            self.sleep(1)
            self.clicks('css', result_sr_data_batch_paste_import_comfirm)
            self.sleep(1)

        # 保存数据
        self.clicks('css', result_save)
        self.wait_loading()
        self.sleep(1)

    # 标准品数据导入
    def standard_sample_import(self):
        """标准品数据导入"""
        log.info("标准品数据导入")
        data = xlrd.open_workbook(wkdl_result_standard_sample_path)
        tables = data.sheets()[0]
        allrows = tables.nrows
        print(allrows)
        num_list = []
        for nss in range(0, allrows - 1):
            h = nss + 1
            vals = tables.row_values(h)
            imp_data = '\t'.join(map(str, vals))
            num_list.append(imp_data)
        print("\n".join(map(str, num_list)))
        pyperclip.copy("\n".join(map(str, num_list)))
        self.clicks('css', result_standard_sample_import_btn)
        self.sleep(0.5)
        self.findelement('css', result_standard_sample_data_batch_paste_import_input).send_keys(Keys.CONTROL, 'v')
        self.sleep(1)
        self.clicks('css', result_standard_sample_data_batch_paste_import_comfirm)
        self.wait_pageinfo_end()
        self.sleep(3)

    def goback_detail(self):
        """返回明细表"""
        urldata = editYaml.read_yaml(functionpageURL_path)
        log.info('点击按钮返回明细表')
        self.click_by_js('css', goback_detail)
        self.sleep(0.5)
        self.clicks('css', goback_page_info)
        self.wait_loading()
        self.sleep(1)

        url = self.get_current_url()  # 获取当前页面URL地址
        print('获取的URL地址', url)
        urldata["url"] = url
        with open(functionpageURL_path, 'w', encoding='utf-8') as fs:  # 写入模式获取的URL地址到yaml文件中
            yaml.safe_dump(urldata, fs, allow_unicode=True)
        print("写入后的URL地址", urldata)

    # 结果表提交
    def result_submit_sample(self):
        """
        结果表提交
        """
        log.info("文库定量结果表提交操作")
        self.clicks('css', result_all_choice)
        self.clicks('css', result_submit)  # 提交按钮
        self.wait_loading()

        # 调用自定义截图方法
        Screenshot(self.driver).get_img("文库定量结果表结果表点击提交按钮","弹出提交确认按钮")

        self.clicks('css', result_submit_comfirm)  # 提交确认按钮
        pageinfo = self.get_pageinfo()
        print(pageinfo)
        self.wait_loading()
        self.wait_pageinfo_end()
        self.sleep(1)
        return pageinfo

    # 结果表样本流程环节写入Excel
    def write_data_to_excel(self):
        """
         根据结果表样本下一步流程，把对应的样本lims号、实验室号、下一步流程以追加形式写入该流程的Excel
        """
        self.add_excel_xlxs(dl_next_step, libquant_result, result_task_id)
        print("下一步流程写入成功")

    # 完成任务单
    def complete_task(self):
        """
        完成任务单
        """

        log.info(' 文库定量结果表,点击完成任务单')

        self.clicks('css', enter_result_list_btn)
        self.wait_loading()
        self.sleep(0.5)

        self.clicks('css', result_complete_task_btn)
        self.wait_loading()

        # 调用自定义截图方法
        Screenshot(self.driver).get_img("文库定量点击完成任务单按钮","完成任务单成功，状态改为完成")

        self.sleep(0.5)

        taskstatus = self.get_text('css', task_status)
        print(taskstatus)
        return taskstatus

    # 文库定量任务列表查询样本所在任务单
    def serach_task(self,path):
        """
        首页面查询已完成的样本任务单
        """
        log.info("首页面查询已完成的样本任务单")
        lims_id = read_excel_col(path, '文库名称')

        self.clicks('css', search)
        self.sleep(1)
        self.input('css', search_lims_sample_num, lims_id[0])
        self.sleep(1)
        self.clicks('css', search_confirm)
        self.wait_loading()
        self.sleep(1)
        samples = self.findelements('xpath', sample_page_list)
        return len(samples)
