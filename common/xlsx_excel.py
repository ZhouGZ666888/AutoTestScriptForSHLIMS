# coding=UTF-8
import openpyxl, os
from xlrd import sheet
from common.all_path import excel_doc_file_path
import pandas as pd


def write_excel_xlsx(path, value1, value2, colname1, colname2):
    """
    把各流程生成的数据写入Excel，需要两个参数，实验流程模块传入lims号和实验室号，订单模块传订单号和患者名称
    :param colname2: 列名2
    :param colname1: 列名1
    :param path: 写入文件的路径
    :param value1:传入find_elements方法获取的多个元素，直接传入即可
    :param value2:find_elements方法获取的多个元素，直接传入eles即可(eles=find_elements(ele))
    :return:
    """
    # 二维list
    company_name_list = []
    value_dict = dict(zip(value1, value2))
    for i, j in value_dict.items():
        print(i.text, j.text)
        company_name_list.append([i.text, j.text])

    df = pd.DataFrame(company_name_list, columns=[colname1, colname2])
    # 保存到本地excel
    df.to_excel(path, index=False)
    print("表格写入数据成功！")


def get_firstDownloadFile():
    """读取系统下载文件的文件位置，按下载时间倒序排列，取最新的一个文件"""
    lists = os.listdir(excel_doc_file_path)
    lists.sort(key=lambda fn: os.path.getmtime(excel_doc_file_path + '\\' + fn))
    filepath = os.path.join(excel_doc_file_path, lists[-1])
    print(filepath)
    return filepath


def get_lims_for_excel(path):
    """
    从对应的Excel中获取上一步流传下来的本节点的待选表lims样本号
    """
    lims_nub = read_excel_col(path, 'lims号')
    lims_id_str = "\n".join(lims_nub)  # 取出Excel表中样本，拼接成字符串录入到检索文本中
    return lims_id_str


def read_excel_col(path, colName):
    """
    根据列名，读取指定Excel的列，输出为list
    :param path: excel文件名
    :param colName: 要读取的列名
    :return: 返回列组成的list，['','','','']
    """
    df = pd.read_excel(path, keep_default_na=False)  # 使用pd读取excel
    df_data = df[colName].values.tolist()
    return df_data


def write_excel_xlsx_by_openpyxl(path, sheet_name, value):
    """
    新建实验流程开始时的样本流转记录表，记录从样本处理开始，各实验流程结果表中样本的下一步流程，openpyxl操作单元格行列下标从（1,1）开始
    :param path: 建立对应下一步流程的Excel表
    :param sheet_name: 记录表sheet页名
    :param value: 要记录的样本数据，以列表嵌套列表格式，子列表数量不限，[[''],[''],[''],...]
    """
    index = len(value)
    workbook = openpyxl.Workbook()  # 新建工作簿（默认有一个sheet）
    sheet_num = workbook.active  # 获得当前活跃的工作页，默认为第一个工作页
    sheet.title = sheet_name + '明细表数据'  # 给sheet页的title赋值
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet_num.cell(row=i + 1, column=j + 1, value=str(value[i][j]))  # 行，列，值 这里是从1开始计数的
    workbook.save(path)  # 一定要保存
    print("xlsx格式表格写入数据成功！")


def pandas_write_excel(company_name_list, path):
    """
    pandas写入Excel
    :param company_name_list:二维列表
    :param path: 文件保存位置
    """
    # list转dataframe
    df = pd.DataFrame(company_name_list)

    # 保存到本地excel，不带表头
    df.to_excel(path, header=None, index=False)


def add_write_excel_xlsx(path, value):
    """
    往各实验流程样本流转记录表中追加数据，对在不同实验流程节点结果表生成的不同的下一步样本，追加到对应的流程记录表中
    :param path: 对应流程记录表
    :param value: 要记录的样本数据，以列表嵌套列表格式，子列表数量不限，[[''],[''],[''],...]
    """
    data = openpyxl.load_workbook(path)
    # 取第一张表
    table = data.sheetnames[0]

    table = data.active
    print(table.title)  # 输出表名
    nrows = table.max_row  # 获得行数
    # 注意行业列下标是从1开始的
    for i in range(1, len(value) + 1):
        for j in range(1, len(value[i - 1]) + 1):
            table.cell(nrows + i, j).value = value[i - 1][j - 1]
    data.save(path)


def padnas_get_column_rows(path, index):
    """
    获取Excel文件去除表头后的总行数或总列数,
    :param path: Excel文件
    :param index: 0-行数，1列数
    :return: 行或列的总数
    """
    df = pd.read_excel(path, header=0)
    return df.shape[index]


def getDataByRowColumn(path, rows, column):
    """
    获取指定单元格的值，行列下标去除标题，从【0,0】开始
    :param path: 文件
    :param rows: 行号
    :param column: 列号
    :return: 单元格值
    """
    data = pd.read_excel(path, header=0)
    ordreNumb = data.iloc[rows, column]
    return str(ordreNumb)


values = [['Es', 'Xs', 'Cs'],
          [7, 8, 9, ],
          ['e', 'f', 'g']]

if __name__ == '__main__':
    c = get_firstDownloadFile()
    print(c)

    # for i in lims_nub1:
    #     # st = lims_nub1[i] + "\n" + lims_nub1[1] + "\n" + lims_nub1[2] + "\n" + lims_nub1[3]
    #     sl.append(i[0])
    #     print(sl)
    # st='\n'.join(sl)
    # print(st)
    # book_name_xlsx = 'xlsx格式测试工作簿.xlsx'

    # sheet_name_xlsx = 'xlsx格式测试表'

    # value3 = [["姓名", "性别", "年龄", "城市", "职业"],
    #           ["111", "女", "66", "石家庄", "运维工程师"],
    #           ["222", "男", "55", "南京", "饭店老板"],
    #           ["333", "女", "27", "苏州", "保安"], ]

    # write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, value3)
    # read_excel_xlsx(book_name_xlsx, sheet_name_xlsx)

    # path = originorder_file_path
    # read_excel_xls(path)
    # test1 = read_excel_xlsx_list_col(path, 0, '订单号')
    # print(test1)
