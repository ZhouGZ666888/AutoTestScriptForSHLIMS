import datetime, xlrd
import re

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait

from common.DataBaseConfig import executeSql
from common.editYaml import read_yaml, save_yaml
from common.xlsx_excel import *

# from pageobj.pathologycheckPage import now_time
from data.sql_action.execute_sql_action import *


def wait_loading():
    """
    设置等待页面loading结束再去操作,是结束，不是出现
    el-message el-message--success
    """
    driver = webdriver.Chrome()
    loading = '//*[@class="el-loading-mask is-fullscreen"]/descendant::p[text()="Loading"]'  # 定义了loading
    # 等待60s超时，默认0.5s寻找一次
    WebDriverWait(driver, 50).until_not(lambda x: x.find_element(By.XPATH, loading))
    time.sleep(1)


# driver.get(r'C:\Users\admin\Desktop\工作文件夹\自动化文档\其他脚本\Lims_Test\common\index.html')
# wait_loading()
# print('jieshu')

sm = """
SELECT
T.sample_id_lims AS sampleIdLims,
T.position_in_box AS positionInBox,
t2.sample_type_name AS sampleTypeName,
t5.box_name AS boxName,
t7.position_code AS positionCode,
t7.storage_name AS storageName,
tv.sample_id_lab AS sampleIdLab,
t8.dt_name_cn AS storageTypeName,
t5.position_in_drawer AS positionInDrawer,
T.pooling_lims_id AS poolingLimsId
FROM
	sample_info_t
	T LEFT JOIN sample_id_lab_v tv ON ( T.sample_id_lims = tv.sample_id_lims )
	LEFT JOIN bas_sample_type_t t2 ON ( T.sample_type = t2.sample_type_id )
	LEFT JOIN bas_dictionary_t t3 ON ( t3.dt_code = T.sample_amt_unit AND t3.dt_type = 'meterage_unit' )
	LEFT JOIN bas_dictionary_t t4 ON ( t4.dt_code = T.sample_pkg_amt_unit AND t4.dt_type = 'packing_unit' )
	LEFT JOIN sample_box_info_t t5 ON ( T.box_id = t5.box_id )
	LEFT JOIN sample_storage_info_t t6 ON ( t6.storage_id = t5.storage_id )
	LEFT JOIN sample_storage_info_t t7 ON ( t6.parent_id = t7.storage_id )
	LEFT JOIN bas_dictionary_t t8 ON ( t7.storage_type = t8.dt_code AND t8.dt_type = 'storage_type' ) 
WHERE
	T.is_valid = '1' 
	
	AND T.sample_status = '01' 
	AND T.sample_pkg_amt >= 0 
	AND (
		( t7.storage_type = '01' ) 
		OR (
			t7.storage_type = '00' 
			AND EXISTS ( SELECT 1 FROM sample_storage_user_t ssu WHERE ssu.is_valid = '1' AND ssu.storage_id = t7.storage_id AND ssu.user_id = 'zhouguanzhong' ) 
		) 
	) 
	AND t7.storage_type = '00' 
	AND t5.box_name = '自动化测试用(勿删)' 
ORDER BY
	T.mod_date DESC 
	LIMIT 10 OFFSET 0;

"""

seach = """
SELECT T.sample_id_lims AS sampleIdLims,T.position_in_box AS positionInBox,t2.sample_type_name AS sampleTypeName,
t5.box_name AS boxName,t7.position_code AS positionCode,t7.storage_name AS storageName,tv.sample_id_lab AS sampleIdLab,t8.dt_name_cn AS storageTypeName,t5.position_in_drawer AS positionInDrawer,T.pooling_lims_id AS poolingLimsId FROM
	sample_info_t
	T LEFT JOIN sample_id_lab_v tv ON ( T.sample_id_lims = tv.sample_id_lims )
	LEFT JOIN bas_sample_type_t t2 ON ( T.sample_type = t2.sample_type_id )
	LEFT JOIN bas_dictionary_t t3 ON ( t3.dt_code = T.sample_amt_unit AND t3.dt_type = 'meterage_unit' )
	LEFT JOIN bas_dictionary_t t4 ON ( t4.dt_code = T.sample_pkg_amt_unit AND t4.dt_type = 'packing_unit' )
	LEFT JOIN sample_box_info_t t5 ON ( T.box_id = t5.box_id )
	LEFT JOIN sample_storage_info_t t6 ON ( t6.storage_id = t5.storage_id )
	LEFT JOIN sample_storage_info_t t7 ON ( t6.parent_id = t7.storage_id )
	LEFT JOIN bas_dictionary_t t8 ON ( t7.storage_type = t8.dt_code AND t8.dt_type = 'storage_type' ) 
WHERE
	T.is_valid = '1' 
		AND  (strpos(t.sample_id_lims, '{}') >= 1 OR strpos(t.sample_id_lims, '{}') >= 1 OR strpos(t.sample_id_lims, '{}') >= 1 OR strpos(t.sample_id_lims, '{}') >= 1 OR strpos(t.sample_id_lims, '{}') >= 1 OR strpos(t.sample_id_lims, '{}') >= 1 OR strpos(t.sample_id_lims, '{}') >= 1 OR strpos(t.sample_id_lims, '{}') >= 1 OR strpos(t.sample_id_lims, '{}') >= 1 OR strpos(t.sample_id_lims, '{}') >= 1) 
	AND T.sample_status = '01' 
	AND T.sample_pkg_amt >= 0 
	AND (
		( t7.storage_type = '01' ) 
		OR (
			t7.storage_type = '00' 
			AND EXISTS ( SELECT 1 FROM sample_storage_user_t ssu WHERE ssu.is_valid = '1' AND ssu.storage_id = t7.storage_id AND ssu.user_id = 'zhouguanzhong' ) 
		) 
	) 
	AND t7.storage_type = '00' 
	AND t5.box_name = '自动化测试用(勿删)' 
ORDER BY
	T.mod_date DESC 
	LIMIT 10 OFFSET 0;
"""

# sql = GetSqlHelper()
# data = execute_sql.test_select_limsdb(yk_sample_info)
# ilist=list(i['sampleidlims']) for i in data
# print(data)
# lims_list = [item['sampleidlims'] for item in data]
# print(lims_list)
# data2 = sql.test_select_limsdb(yk_sample_search.format(lims_list[0],lims_list[1],lims_list[2],lims_list[3],
#                                                       lims_list[4],lims_list[5],lims_list[6],lims_list[7],lims_list[8],lims_list[9]))
# lims_list2 = [item['sampleidlims'] for item in data2]
# print(lims_list2)
# now = time.strftime("%Y-%m-%d-%H:%M:%S")
# print(now)
from common.all_path import *

sqldata = "SELECT box_name FROM sample_box_info_t t WHERE t.box_name LIKE '盒子_20%'   ORDER BY creation_date DESC;"
# print(sqldata)
# ret = execute_sql.test_select_limsdb(sqldata)
# # print(ret)
# print(ret)
box_name = '盒子_2023-02-22-19:31:08'
sqlData = "SELECT task_id FROM sample_transfer_item_t t1 LEFT JOIN sample_box_info_t t2 ON (" \
          "t1.box_id=t2.box_id) WHERE t2.box_name  IN (" + "'" + box_name + "'" + ")"
ybcl_detail_sql = "UPDATE exp_preparation_item_t set position_in_box='1' where task_id='{}';"


# from common.xlsx_excel import padnas_get_column_rows
# sql_first_step.test_updateByParam(ybcl_detail_sql.format('YBCL2023022300001'))
# import pandas as pd
# from common.xlsx_excel import *
#
# from datetime import datetime
# sr_sample_id_external = now_time.strftime('%Y%m%d%H%M') + '_TEST_SR'
# print(sr_sample_id_external)
# df=pd.read_excel("vdvdv.xlsx",header=0)
# df['组别']='34'
# df.to_excel("vdvdv.xlsx", index=False)
# # df=padnas_get_column_rows(twentyonegene_file_path,0)
# print(df )
# filepath = get_firstDownloadFile()
#
# data = pd.read_excel(order_file_path, header=0)
# s=data.iloc[0,0]
# data["组别"] = 5
# data["复溶体积μL"] = "22"
# data["梯度时间mi"] = "5"
# data["上机序列号"] = 'SNJ' + datetime.now().strftime('%Y%m%d%H%M')
# data["项目编号"] = 'J022'
#
# data.to_excel(filepath, index=False)

# print(type(str(s)))
# sb=['GS2112270002', 'GS2112270003']
# sample_nub = tuple(sb)
# # print(sample_nub)
# new_ordernum=99991412
# dic1 = {'order_number': [str(new_ordernum)],
#         'patient_name': ['崔东山']
#         }
# df = pd.DataFrame(dic1, dtype=object)
# df.to_excel(order_file_path, index=False)
# order_nub = read_excel_col(order_file_path, 'order_number')  # 订单Excel获取订单号
# enrichment_nub = read_excel_col(wkdl_sr_file_path, 'lims号')
# print(enrichment_nub)
# datetime.datetime.now()
# sd= str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))[:]
# print(sd)

# datetime2=datetime.datetime.now()+datetime.timedelta(minutes=10)
# print(str((datetime.datetime.now()+datetime.timedelta(minutes=10)).strftime('%Y-%m-%d %H:%M:%S'))[:])

# str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))[:]
# tlist=[]
# item={}
# for i in range(0,10):
#
#     item['a']=i
#     item['b']=i+1
#     tlist.append(item)
# print(tlist)
# s=[{'a': 0, 'b': 1}, {'a': 1, 'b': 2}, {'a': 2, 'b': 3}, {'a': 3, 'b': 4}, {'a': 4, 'b': 5}, {'a': 5, 'b': 6},
#   {'a': 6, 'b': 7}, {'a': 7, 'b': 8}, {'a': 8, 'b': 9}, {'a': 9, 'b': 10}]
# data = xlrd.open_workbook(twentyonegene_file_path)
# num_list = []
# for index in range(0, 22):
#     tables = data.sheets()[0]
#     vals = tables.row_values(index)
#     imp_data = '\t'.join(map(str, vals))
#     num_list.append(imp_data)
# print("\n".join(map(str, num_list)))

# detail_list= [
#     {
#         "source_field_name": "机构名称",
#         "source_field": "COMP_NAME"
#     },
#     {
#         "source_field_name": "统一社会信用代码",
#         "source_field": "CREDIT_CODE"
#     }
# ]
# for i in detail_list:
#     print(i['source_field_name'])

class ABC:
    S = 1234534534535345

    def tester(self):
        s = self.S
        print(s)


class Student:
    score = 99.95

    def __init__(self):
        pass

    def test(self):
        print(self.score)


# dic1 = {'order_number':'fdfdsdgdfgd423423423',
#                         'patient_name': '张三'
#                         }
# oredr=read_yaml(orderNub_path)
# print(oredr['order_number'])
# if oredr['order_number'] is None:
#
#
#     print('kooo')
# else:
#     print(';fdsfds')
def get_data(max):
    a = 0
    while a < max:
        yield a
        a += 1


def my_openpyxl(sheet):
    path = csps_file_path
    wb = load_workbook(path)
    wt = wb[sheet]
    test_data = []
    for x in range(2, len(tuple(wt.rows)) + 1):
        data = []
        for y in range(2, 7):
            data.append(wt.cell(row=x, column=y).value)
        test_data.append(data)
    return test_data


key_to_find = "properties"
key_to_find2 = 'sms_code'


def find_key1(key_to_find, nested_json):
    item1 = {}
    if isinstance(nested_json, list):
        for item in nested_json:
            result = find_key1(key_to_find, item)
            if result:
                print(result)
                for k, v in result.items():
                    item1[k] = v['type']
                return item1
    elif isinstance(nested_json, dict):
        for key, value in nested_json.items():
            if key == key_to_find:
                for k, v in value.items():
                    item1[k] = v['type']
                return item1
                # return value
            elif isinstance(value, (dict, list)):
                result = find_key1(key_to_find, value)
                if result:
                    print('this is ', result)
                    for k, v in result.items():
                        item1[k] = v['type']
                    return item1


def find_key(key_to_find, nested_json):
    '''

    :param key_to_find: 要找的键名
    :param nested_json: json串
    :return: 返回key_to_find对应的值
    '''
    if isinstance(nested_json, list):
        for item in nested_json:
            result = find_key(key_to_find, item)
            if result:
                return result
    elif isinstance(nested_json, dict):
        for key, value in nested_json.items():
            if key == key_to_find:
                return value
            elif isinstance(value, (dict, list)):
                result = find_key(key_to_find, value)
                if result:
                    print('this is ', result)
                    return result


def find_key_new(key, json_obj):
    """

    """
    result = []
    if isinstance(json_obj, dict):
        for k, v in json_obj.items():
            if k == key:
                result.append(v)
            else:
                result.extend(find_key_new(key, v))
    elif isinstance(json_obj, list):
        for item in json_obj:
            result.extend(find_key_new(key, item))
    # elif isinstance(json_obj, str):
    #     pass
    return result


testdata = read_yaml(sampledata_path)
order = read_yaml(orderNub_path)  # 获取订单号
import pandas as pd


def wer():
    df = pd.read_csv(mutation_file_path, encoding='utf-8', )
    print(df.loc[:, 'id'])


def sr_sample_import():
    """
    准备sr样本数据，在sr信息登记模块使用。选取一条sr样本，设置sr样本的外部样本编号，并存入对应的导入模板
    """

    now_time = datetime.datetime.now()
    sr_sample_id_external = now_time.strftime('%Y%m%d%H%M') + '_TEST_SR'  # 按时间规则生成外部样本编号

    wb = load_workbook(filename=sr_sample_imp_file)  # 打开excel文件
    ws = wb.active
    ws.cell(2, 2, sr_sample_id_external)  # 修改第k行，第index列值
    wb.save(sr_sample_imp_file)

    wb = load_workbook(filename=sr_sample_sublibrary_imp_file)  # 打开excel文件
    ws = wb.active
    ws.cell(2, 1, sr_sample_id_external)  # 修改第k行，第index列值
    wb.save(sr_sample_sublibrary_imp_file)

    # 数据库先获取sr样本的lims号，再根据lims号在数据库设置外部样本编号
    sr_sample_nubs = executeSql.test_select_limsdb(get_sr_sample_lims.format(order['order_number']))
    sr_sample_nub = [list(i.values()) for i in sr_sample_nubs]
    print('选中的SR样本：', sr_sample_nub[0][0])
    executeSql.test_updateByParam(set_sr_sample_id_external.format(sr_sample_id_external, sr_sample_nub[0][0]))

    # 将修改后的sr样本的lims号，存到临时文件，在SR样本信息登记模块使用
    datas = read_yaml(sampledata_path)
    datas["rec_sr_sample_for_sr_import"] = sr_sample_nub[0][0]
    save_yaml(sampledata_path, datas)


def test1():
    sr_sample = executeSql.test_select_limsdb(get_sr_sample_lims.format('99991489'))
    sr_sampleLims = [list(i.values()) for i in sr_sample]
    print('选中的SR样本：', sr_sampleLims)
    lims = []
    for i in range(len(sr_sampleLims)):
        sr_sample_id_external = sr_sampleLims[i][0] + '_TEST_SR'
        print(sr_sample_id_external)

        wb = load_workbook(filename=sr_sample_imp_file)  # 打开excel文件
        ws = wb.active
        ws.cell(2 + i, 2, sr_sample_id_external)  # 修改第k行，第index列值
        wb.save(sr_sample_imp_file)

        wb = load_workbook(filename=sr_sample_sublibrary_imp_file)  # 打开excel文件
        ws = wb.active
        ws.cell(2 + i, 1, sr_sample_id_external)  # 修改第k行，第index列值
        wb.save(sr_sample_sublibrary_imp_file)
        executeSql.test_updateByParam(set_sr_sample_id_external.format(sr_sample_id_external, sr_sampleLims[i][0]))
        lims.append(sr_sampleLims[i][0])
    datas = read_yaml(SR_sample_for_import_path)
    datas["rec_sr_sample_for_sr_import"] = lims
    save_yaml(SR_sample_for_import_path, datas)


from itertools import product


def eee():
    lims_id = executeSql.test_select_limsdb(
        app_get_result_lims.format('APPA2023082300008'))  # 从数据库获取当前任务单号下样本lims号
    # 把获取的lims号转换为一维列表
    blist = [[item[i] for i in item] for item in lims_id]
    print(blist)
    list1 = [5, 3, 4, 7, 5, 4, 32]  # 批量导入文库投入量、实际取样体积、补Buffer体积、PCR循环数、产物浓度、产物体积、产物总量
    impData = []
    for i in blist:
        new_list = i + list1
        impData.append(new_list)
    pandas_write_excel(impData, position_in_box_path)  # 把样本号和盒内位置编号写入Excel模板


def redf():
    data = [['GS2307210003J001', 'KA237L0003-L000J006-APP-A-1'],
            ['GS2307210003J004', 'KA237L0003-L000J006-APP-A-4'],
            ['GS2307210003J007', 'KA237L0003-L000J006-APP-A-7']]

    # Create a DataFrame
    df = pd.DataFrame(data)

    # Write to Excel
    df.to_excel("my_data.xlsx", index=False, header=False)

    # Read from Excel
    new_df = pd.read_excel("my_data.xlsx", header=None)
    # for i in new_df:
    #     imp_data = '\t'.join(map(str, i))
    print(new_df)


def test22():
    # enrichment_nub = read_excel_col(wkdl_sr_file_path, 'lims号')
    enrichment_nub = read_excel_col(wkdl_sr_file_path, 'lims号')
    print(enrichment_nub)
    if enrichment_nub:
        sample_datas = str(enrichment_nub).replace('[', '').replace(']', '')  # 把取出的列表转换成sql条件中的元组格式
        # 根据接样lims号，到数据库中（sample_info_t表）获取定量富集lims号
        sa = executeSql.test_select_limsdb(
            "SELECT DISTINCT (pooling_lims_id) FROM sample_info_t t1 WHERE t1.original_sample_id_lims IN ({})AND "
            "t1.sr_type = '02'".format(sample_datas))
        print(sa)
        result_sql = [item[key] for item in sa for key in item]  # sql返回值是列表套字典，这里取出字典的值
        while None in result_sql:  # 取出字典值中的空值None
            result_sql.remove(None)
        print('样本接收定量流程SR样本', result_sql)
        return result_sql
    lims_id = executeSql.test_select_limsdb(wkdl_result_sql1.format('WKDL2023082400003'))  # 从数据库获取当前任务单号下样本总数
    lims_count = [item[key] for item in lims_id for key in item]
    sr_samples = read_excel_col(wkdl_hdsr_file_path, 'lims号')
    print(enrichment_nub)


def read_sr_import_data():
    """
    根据sr样本数量，修改sr类型导入模板
    """
    sr_samples = read_excel_col(wkdl_hdsr_file_path, 'lims号')

    wb = load_workbook(filename=wkdl_detail_sr_import_path)  # 打开excel文件
    # 把流程中的样本号写入待导入模板
    for i in range(0, len(sr_samples)):
        k = i + 2

        ws = wb.active
        # 根据需要修改表格数据
        ws.cell(k, 1, sr_samples[i])  # 修改第k行，第index列值
    wb.save(wkdl_detail_sr_import_path)

    # 从模板中复制出修改后的待导入数据
    data = xlrd.open_workbook(wkdl_detail_sr_import_path)
    num_list = []
    for index in range(0, len(sr_samples)):
        h = index + 1
        tables = data.sheets()[0]
        # allrows = tables.nrows
        vals = tables.row_values(h)
        imp_data = '\t'.join(map(str, vals))
        num_list.append(imp_data)
    print("\n".join(map(str, num_list)))


def get_non_sr_sample_from_excel(col_name):
    """
     从对应的Excel中获取上一步流传下来的本节点的待选表lims样本号
    :param col_name: excel列名
    :return:lims_nub
    """

    lims_nub = read_excel_col(wkdl_non_sr_file_path, col_name)
    print(lims_nub)
    if lims_nub:
        return lims_nub
    else:
        return None


import re
from bs4 import BeautifulSoup

html = r'C:\Users\admin\Desktop\html1.html'


def extract_chinese(html):
    soup = BeautifulSoup(html, 'html.parser')
    text = soup.get_text()
    chinese = re.findall(r'[\u4e00-\u9fff]+', text)
    return chinese


def test111():
    sr_sampleLims = [['GS2309050069'], ['GS2309050071'], ['GS2309050072']]
    sr_sample_imp_data = [['J022'], ['靶向富集'], ['HiseqX'], ['PE150'], ['单'], ['不混样包Lane'], [1], ['G'], ['单梯度绝对'],
                          [2100], [0.796],
                          [2.3], [24], [330], ['双标签'], ['是'], ['是'], [2], ['否'], [12], ['暂无'], [2], [34], ['包埋文库'],
                          ['F类'],
                          ['备注'], [datetime.datetime.now().strftime('%Y.%m.%d')], [1], [2], ['否']]

    sr_sample_sublibrary_imp_data = [['O01-004'], ['DC-013'], ['AGTCT'], ['DC-014'], ['DC-015'], ['J022']]
    sr_sample_imp_file_r = load_workbook(sr_sample_imp_file)
    sr_sample_sublibrary_imp_file_r = load_workbook(sr_sample_sublibrary_imp_file)
    # 取第一张表

    lims = []
    for val in range(len(sr_sampleLims)):

        sample_id_external = [sr_sampleLims[val][0] + '_TEST_SR']
        lims.append(sr_sampleLims[val][0])

        sr_sample_imp_data.insert(1, sample_id_external)
        sr_sample_sublibrary_imp_data.insert(0, sample_id_external)

        table = sr_sample_imp_file_r.active
        table1 = sr_sample_sublibrary_imp_file_r.active
        nrows = table.max_row  # 获得行数
        nrows1 = table1.max_row  # 获得行数
        # 注意行业列下标是从1开始的
        for i in range(1, len(sr_sample_imp_data) + 1):
            for j in range(1, len(sr_sample_imp_data[i - 1]) + 1):
                table.cell(nrows + 1, i).value = sr_sample_imp_data[i - 1][j - 1]
        sr_sample_imp_data.pop(1)

        for i in range(1, len(sr_sample_sublibrary_imp_data) + 1):
            for j in range(1, len(sr_sample_sublibrary_imp_data[i - 1]) + 1):
                table1.cell(nrows1 + 1, i).value = sr_sample_sublibrary_imp_data[i - 1][j - 1]
        sr_sample_sublibrary_imp_data.pop(0)
        executeSql.test_updateByParam(set_sr_sample_id_external.format(sample_id_external[0], sr_sampleLims[val][0]))
    sr_sample_imp_file_r.save(sr_sample_imp_file)
    sr_sample_sublibrary_imp_file_r.save(sr_sample_sublibrary_imp_file)
    datas = read_yaml(SR_sample_for_import_path)
    datas["rec_sr_sample_for_sr_import"] = lims
    save_yaml(SR_sample_for_import_path, datas)


def test22ss():
    s='KA239A0007-XXXXXXBXDF03F5-J022'
    print(s.strip()[:11])

if __name__ == '__main__':
    # html = r'C:\Users\admin\Desktop\html1.html'
    test22ss()

    # nested_json=open('cstest.json','r',encoding='utf8')
    # data = json.load(nested_json)
    # nested_json.close()
    # print(data)
    # result = find_key_new(key_to_find, data)
    # print(result)
    # result2 = find_key(key_to_find, data['paths'])
    # print('result2',result2)

#
# s={'phone_number': {'type': 'string', 'title': '界面输入-手机号'}, 'sms_code': {'type': 'string', 'title': '界面输入-短信码'}}
# result = find_key(key_to_find, data['paths'])
# item={}
# for  k,v in result.items():
#     # print(k,v)
#     item[k]=v['type']
# print(item)
